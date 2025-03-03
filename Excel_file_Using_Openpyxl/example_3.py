# import openpyxl module 
import openpyxl

# Call a Workbook() function of openpyxl  
# to create a new blank Workbook object 
wb = openpyxl.Workbook()

# Get workbook active sheet   
# from the active attribute 
sheet = wb.active
# wb.create_sheet(index=2 , title="demo sheet2")
# wb.create_sheet(index=3 , title="demo sheet3")
wb.create_sheet(index=4 , title="demo sheet4")

wb.save("D:\Python_Projects_for_Beginners\Excel_file_Using_Openpyxl\demo.xlsx")
print("Saved !!")