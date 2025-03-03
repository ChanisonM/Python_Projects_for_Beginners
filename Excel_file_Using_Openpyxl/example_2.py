# import openpyxl module 
import openpyxl

# Call a Workbook() function of openpyxl  
# to create a new blank Workbook object 
wb = openpyxl.Workbook()

# Get workbook active sheet   
# from the active attribute 
sheet = wb.active

# Cell objects also have row, column 
# and coordinate attributes that provide 
# location information for the cell. 
  
# Note: The first row or column integer 
# is 1, not 0. Cell object is created by 
# using sheet object's cell() method.
c1 = sheet.cell(row=1 , column=1)
# writing values to cells
c1.value = "ANKIT"

c2 = sheet.cell(row=1 ,column=2)
c2.value = "RAI"

c3 = sheet.cell(row=3,column=1)
c3.value = "Chanison"

c4 = sheet.cell(row=3,column=2)
c4.value = "Aupathin"



# Anytime you modify the Workbook object 
# or its sheets and cells, the spreadsheet 
# file will not be saved until you call 
# the save() workbook method. 

wb.save("D:\Python_Projects_for_Beginners\Excel_file_Using_Openpyxl\demo.xlsx")
print("Saved !!")